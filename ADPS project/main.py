#Python inbuild libraries :
import datetime
import random

#External libraries :
#Tkinter for GUI :
import tkinter as tk

#Pathilib to find file path
import pathlib

#For working with excel files :
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

#imported library for creating excel :
import creating_excel as xl

xl.creating_excel()


parking_file=pathlib.Path.cwd()/"parking_details.xlsx"
user_file=pathlib.Path.cwd()/"user_log.xlsx"

wb=load_workbook(filename=user_file)
wb2=load_workbook(filename=parking_file)
sheet=wb.active
sheet2=wb2.active


#constant
BGCOLOR = "#020f12"
CYAN = "#05d7ff"
LIGHTCYAN= "#65e7ff"
BLACK = "BLACK"
WHITE ="WHITE"


# WINDOW :
window = tk.Tk()
window.title("ADPS PARKING")
window.geometry("1920x1080")
#window.resizable(False,False)
window.configure(background=BGCOLOR)


#ICON :
logo_file=pathlib.Path.cwd()/"logo.png"
icon_image=tk.PhotoImage(file=logo_file)
window.iconphoto(False,icon_image)


def slots_availability():

    global available_slots
    global total_slots
    global total_available_slots

    dataframe=pd.read_excel(parking_file)

    available_slots=dataframe['STATUS'].tolist()
    total_slots = len(available_slots)
    total_available_slots=available_slots.count(True)


# MAIN PAGE :
def main_page():

    slots_availability()
#---------------------------------------------------------------------------------------------------------------------------------------
# ENTRY PAGE CODE STARTS FROM HERE :
#----------------------------------------------------------------------------------------------------------------------------------------
    #Page after clicking the entry button :
    def entry_page():
        
        page_title.destroy()
        page_subtitle.destroy()
        page_msg.destroy()
        entry_button.destroy()
        exit_button.destroy()

        if total_available_slots !=0:
                
                entrypage_title=tk.Label(window, 
                                bg=CYAN, 
                                fg=BLACK, 
                                text="ENTRY PROCESS :",
                                font=('georgia',90, "bold")
                                )
                entrypage_title.pack(padx=0, pady=45)
                
                entrypage_msg = tk.Label(window, 
                                background=BGCOLOR,
                                foreground=LIGHTCYAN,
                                text = "Enter the Following Details : ",
                                font=('georgia',35)
                                )
                entrypage_msg.place(x=450,y=255)

                mobileno_title = tk.Label(window, 
                                background=BGCOLOR,
                                foreground=WHITE,
                                text = "Mobile number : (OTP will be sent to this number)",
                                font=('georgia',25)
                                )
                mobileno_title.place(x=200,y=350)      

                vehicleno_title = tk.Label(window, 
                                background=BGCOLOR,
                                foreground=WHITE,
                                text = "Vehicle number : ",
                                font=('georgia',25)
                                )
                vehicleno_title.place(x=200,y=545)


                #DATA ENTRY FOR ENTRY :

                mobile_number=tk.StringVar()
                vehicle_number=tk.StringVar()

                mobile_number_entry = tk.Entry(window,
                                                textvariable=mobile_number,
                                                width=21,
                                                border=2,
                                                font=("arial",40)
                                                )
                mobile_number_entry.place(x=300,y=425,height=60)

                vehicle_number_entry = tk.Entry(window,
                                                textvariable=vehicle_number,
                                                width=21,
                                                border=2,
                                                font=("arial",40)
                                                )
                vehicle_number_entry.place(x=300,y=620,height=60)



        #ENTRY END PAGE :
                def entry_end():


                    entrypage_title.destroy()
                    entrypage_msg.destroy()
                    submit_button.destroy()
                    clear_button.destroy()
                    back_button.destroy()
                    mobileno_title.destroy()
                    vehicleno_title.destroy()
                    mobile_number_entry.destroy()
                    vehicle_number_entry.destroy()


                    entered_mobileno=mobile_number.get()
                    entered_vehicleno=vehicle_number.get().upper()
        

                    date_time = datetime.datetime.now()
                    current_date = date_time.strftime("%d-%m-%y")
                    entry_time = date_time.strftime('%H:%M:%S')

                    random_otp=random.randint(100000,999999)


                    if entered_mobileno!=entered_vehicleno:
                            
                            if  entered_mobileno.isdigit() and len(entered_mobileno)== 10:
                                    
                                    for i in range(1,total_slots+1):

                                                    if available_slots[i-1]==True:

                                                        location=sheet2[f'C{i+1}']
                                                        
                                                        sheet2[f'B{i+1}']='False' #pandas
                                                        sheet2[f'D{i+1}']=int(random_otp) #pandas
                                                        sheet2[f'E{i+1}']=entered_vehicleno #pandas


                                                        sheet.cell(column=1,row=sheet.max_row+1,value=current_date)
                                                        sheet.cell(column=2,row=sheet.max_row,value=entry_time)
                                                        sheet.cell(column=3,row=sheet.max_row,value=i)
                                                        sheet.cell(column=4,row=sheet.max_row,value=int(entered_mobileno))
                                                        sheet.cell(column=5,row=sheet.max_row,value=entered_vehicleno)
                                                        sheet.cell(column=6,row=sheet.max_row,value=random_otp)
                                                        break


                                    entryend_title=tk.Label(window, 
                                                bg=CYAN, 
                                                fg=BLACK, 
                                                text="COMPLETED ✔️",
                                                font=('georgia',90, "bold")
                                                )
                                    entryend_title.place(x=280,y=45)

                                    entryend_msg=tk.Label(window,
                                                background=BGCOLOR,
                                                foreground=WHITE,
                                                text='''
Your entry process is completed successfully, And your receipt has 
been sent to your moblie number. You can complete the payment process
at the exit gate .
Thankyou Have a Nice Day !

\"OTP is generated and sent your mobile number\"'''     ,
                                                font=('georgia',27)
                                                )
                                    entryend_msg.place(x=180,y=230)

                                    entryend_slotno=tk.Label(window,
                                                background=BGCOLOR,
                                                foreground=WHITE,
                                                text=(f"Your booked slot no is : {i} "),
                                                font=('georgia',35)
                                                )
                                    entryend_slotno.place(x=500,y=550)

                                    def confirm():
                                            confirm_button.destroy()
                                            entryend_title.destroy()
                                            entryend_msg.destroy()
                                            entryend_slotno.destroy()
                                                
                                            main_page()

                                    confirm_button=tk.Button(window, 
                                            background=CYAN,
                                            foreground=BLACK,
                                            activebackground=LIGHTCYAN,
                                            activeforeground=BLACK,
                                            highlightthickness=2,
                                            highlightbackground=CYAN,
                                            highlightcolor=WHITE,
                                            width=9,
                                            height=1,
                                            border=0,
                                            cursor='hand1',
                                            text="CONFIRM",
                                            font=('arial',40,'bold'),
                                            command=confirm
                                            )
                                    confirm_button.place(x=620,y=650)
                                    
                                    wb.save(user_file)
                                    wb2.save(parking_file)
                            
                            
                            else:
                             # Handle the case where entered_otp_number is not a valid integer
                             # You can display an error message to the user or take any other appropriate action
                                    wrong_exit_input_title=tk.Label(window, 
                                                    bg=CYAN, 
                                                    fg=BLACK, 
                                                    text="WRONG ❌",
                                                    font=('georgia',90, "bold")
                                                    )
                                    wrong_exit_input_title.place(x=400,y=45)

                                    wrong_exit_input_msg=tk.Label(window,
                                                    background=BGCOLOR,
                                                    foreground=WHITE,
                                                    text='''
The OTP or the slot number that your entered is \"INVALID\"
Please enter correct info to complete the exit procedure.

Press the retry button below to return to the previous page!!!

TRY AGAIN !!!'''     ,
                                                    font=('georgia',33)
                                                    )
                                    wrong_exit_input_msg.place(x=175,y=230)

                                    def retry():
                                                retry_button.destroy()
                                                wrong_exit_input_title.destroy()
                                                wrong_exit_input_msg.destroy()
                                                entry_page()

                                    retry_button=tk.Button(window, 
                                                    background=CYAN,
                                                    foreground=BLACK,
                                                    activebackground=LIGHTCYAN,
                                                    activeforeground=BLACK,
                                                    highlightthickness=2,
                                                    highlightbackground=CYAN,
                                                    highlightcolor=WHITE,
                                                    width=9,
                                                    height=1,
                                                    border=0,
                                                    cursor='hand1',
                                                    text="Retry",
                                                    font=('arial',40,'bold'),
                                                    command=retry
                                                    )
                                    retry_button.place(x=620,y=650)
                    
                        
                    elif entered_mobileno==entered_vehicleno:
                             
                                emptyentry_title=tk.Label(window, 
                                        bg=CYAN, 
                                        fg=BLACK, 
                                        text="WORNG ❌",
                                        font=('georgia',90, "bold")
                                        )
                                emptyentry_title.place(x=400,y=45)

                                emptyentry_msg=tk.Label(window,
                                                background=BGCOLOR,
                                                foreground=WHITE,
                                                text='''
The mobile number or the vehicle number that your entered is \"INVALID\"
or \"EMPTY\"Please enter correct info to complete the entry procedure.

Press the retry button below to return to the previous page!!!

TRY AGAIN !!!''' ,
                                                font=('georgia',30)
                                                )
                                emptyentry_msg.place(x=100,y=250)
                             
                                def confirm():
                                        retry_button.destroy()
                                        emptyentry_title.destroy()
                                        emptyentry_msg.destroy()
                                        entry_page()

                                retry_button=tk.Button(window, 
                                        background=CYAN,
                                        foreground=BLACK,
                                        activebackground=LIGHTCYAN,
                                        activeforeground=BLACK,
                                        highlightthickness=2,
                                        highlightbackground=CYAN,
                                        highlightcolor=WHITE,
                                        width=9,
                                        height=1,
                                        border=0,
                                        cursor='hand1',
                                        text="RETRY",
                                        font=('arial',40,'bold'),
                                        command=confirm
                                        )
                                retry_button.place(x=595,y=650)


                    wb.save(user_file)
                    wb2.save(parking_file)
                    



        # BUTTONS FOR ENTRY PAGE STARTS HERE :
                    
                submit_button=tk.Button(window, 
                                background=CYAN,
                                foreground=BLACK,
                                activebackground=LIGHTCYAN,
                                activeforeground=BLACK,
                                highlightthickness=2,
                                highlightbackground=CYAN,
                                highlightcolor=WHITE,
                                width=8,
                                height=1,
                                border=0,
                                cursor='hand1',
                                text="SUBMIT",
                                font=('arial',40,'bold'),
                                command=entry_end,
                                )
                submit_button.place(x=1150,y=350)

                def clear():
                    mobile_number.set('')
                    vehicle_number.set('')

                clear_button=tk.Button(window, 
                                background=CYAN,
                                foreground=BLACK,
                                activebackground=LIGHTCYAN,
                                activeforeground=BLACK,
                                highlightthickness=2,
                                highlightbackground=CYAN,
                                highlightcolor=WHITE,
                                width=8,
                                height=1,
                                border=0,
                                cursor='hand1',
                                text="CLEAR",
                                font=('arial',40,'bold'),
                                command=clear
                                )
                clear_button.place(x=1150,y=500)

                def back():
                    entrypage_title.destroy()
                    entrypage_msg.destroy()
                    submit_button.destroy()
                    clear_button.destroy()
                    back_button.destroy()
                    vehicleno_title.destroy()
                    mobileno_title.destroy()
                    vehicle_number_entry.destroy()
                    mobile_number_entry.destroy()
                    main_page()

                back_button=tk.Button(window, 
                                background=CYAN,
                                foreground=BLACK,
                                activebackground=LIGHTCYAN,
                                activeforeground=BLACK,
                                highlightthickness=2,
                                highlightbackground=CYAN,
                                highlightcolor=WHITE,
                                width=8,
                                height=1,
                                border=0,
                                cursor='hand1',
                                text="BACK",
                                font=('arial',40,'bold'),
                                command=back
                                )
                back_button.place(x=1150,y=650)


        elif total_available_slots==0:

                    noslot_title=tk.Label(window, 
                                bg=CYAN, 
                                fg=BLACK, 
                                text="SORRY ❌",
                                font=('georgia',90, "bold")
                                )
                    noslot_title.place(x=450,y=45)

                    noslot_msg=tk.Label(window,
                                        background=BGCOLOR,
                                        foreground=WHITE,
                                        text='''
Unfortunately, the action you requested cannot be done as there are now 
\"NO EMPTY SLOTS\". Next time, please check the billboard before entering. 

Sorry for the inconvenience, and thank you for understanding.

Thank you. Have a nice day. !'''     ,
                                        font=('georgia',27)
                                        )
                    noslot_msg.place(x=150,y=250)

                    def confirm():
                        confirm_button.destroy()
                        noslot_title.destroy()
                        noslot_msg.destroy()
                        main_page()

                    confirm_button=tk.Button(window, 
                                background=CYAN,
                                foreground=BLACK,
                                activebackground=LIGHTCYAN,
                                activeforeground=BLACK,
                                highlightthickness=2,
                                highlightbackground=CYAN,
                                highlightcolor=WHITE,
                                width=9,
                                height=1,
                                border=0,
                                cursor='hand1',
                                text="CONFIRM",
                                font=('arial',40,'bold'),
                                command=confirm
                                )
                    confirm_button.place(x=620,y=650)
   
#---------------------------------------------------------------------------------------------------------------------------------------
# EXIT PAGE CODE STARTS FROM HERE :
#----------------------------------------------------------------------------------------------------------------------------------------       

    def exit_page():

        dataframe=pd.read_excel(parking_file)

        available_slots=dataframe['STATUS'].tolist()
        total_slots = len(available_slots)
        total_available_slots=available_slots.count(True)
        
        #color_box.destroy()
        page_title.destroy()
        page_subtitle.destroy()
        page_msg.destroy()
        entry_button.destroy()
        exit_button.destroy()

        if total_available_slots!=total_slots:

            exitpage_title=tk.Label(window, 
                            bg=CYAN, 
                            fg=BLACK, 
                            text=" EXIT PROCESS :",
                            font=('georgia',90, "bold")
                            )
            exitpage_title.pack(padx=0, pady=45)
            
            exitpage_msg = tk.Label(window, 
                            background=BGCOLOR,
                            foreground=LIGHTCYAN,
                            text = "Enter the Following Details : ",
                            font=('georgia',35)
                            )
            exitpage_msg.place(x=450,y=255)

            slotno_title = tk.Label(window, 
                            background=BGCOLOR,
                            foreground=WHITE,
                            text = "Slot number : ",
                            font=('georgia',25)
                            )
            slotno_title.place(x=200,y=350)      

            otp_title = tk.Label(window, 
                            background=BGCOLOR,
                            foreground=WHITE,
                            text = "OTP number : ",
                            font=('georgia',25)
                            )
            otp_title.place(x=200,y=545)


            #DATA ENTRY FOR EXIT  :
            
            slot_number=tk.StringVar()
            otp_number=tk.StringVar()

            slot_number_entry = tk.Entry(window,
                                            textvariable=slot_number,
                                            width=21,
                                            border=2,
                                            font=("arial",40)
                                            )
            slot_number_entry.place(x=300,y=425,height=60)

            otp_number_entry = tk.Entry(window,
                                            textvariable=otp_number,
                                            width=21,
                                            border=2,
                                            font=("arial",40)
                                            )
            otp_number_entry.place(x=300,y=620,height=60)

        
            def exit_end():
                
                exitpage_title.destroy()
                exitpage_msg.destroy()
                submit_button.destroy()
                clear_button.destroy()
                back_button.destroy()
                slotno_title.destroy()
                otp_title.destroy()
                otp_number_entry.destroy()
                slot_number_entry.destroy()

                entered_slot_number=slot_number.get()
                entered_otp_number=otp_number.get()
                
                     
                if entered_otp_number=='' or entered_slot_number=='':

                    wrong_exit_input_title=tk.Label(window, 
                                            bg=CYAN, 
                                            fg=BLACK, 
                                            text="WRONG ❌",
                                            font=('georgia',90, "bold")
                                            )
                    wrong_exit_input_title.place(x=400,y=45)

                    wrong_exit_input_msg=tk.Label(window,
                                                    background=BGCOLOR,
                                                    foreground=WHITE,
                                                    text='''
The OTP or the slot number that your entered is \" EMPTY \"
Please enter correct info to complete the exit procedure.

Press the retry button below to return to the previous page!!!

TRY AGAIN !!!'''     ,
                                                    font=('georgia',33)
                                                    )
                    wrong_exit_input_msg.place(x=175,y=230)

                    def retry():
                                    retry_button.destroy()
                                    wrong_exit_input_title.destroy()
                                    wrong_exit_input_msg.destroy()
                                    exit_page()

                    retry_button=tk.Button(window, 
                                            background=CYAN,
                                            foreground=BLACK,
                                            activebackground=LIGHTCYAN,
                                            activeforeground=BLACK,
                                            highlightthickness=2,
                                            highlightbackground=CYAN,
                                            highlightcolor=WHITE,
                                            width=9,
                                            height=1,
                                            border=0,
                                            cursor='hand1',
                                            text="Retry",
                                            font=('arial',40,'bold'),
                                            command=retry
                                            )
                    retry_button.place(x=620,y=650)


                else :
                    
                    
                    try:
                            entered_otp_number2=int(entered_otp_number)
                            entered_slot_number2=int(entered_slot_number)
                    # Your code to handle the integer value goes here

                            check_slot_number = sheet2[f'B{entered_slot_number2+1}'].value
                            check_otp_number= sheet2[f'D{entered_slot_number2+1}'].value


                            if check_slot_number=='False' and entered_otp_number2==check_otp_number:

                # FOR Exit time and total hours calculations :
                        
                                        date_time = datetime.datetime.now()
                                        #current_date = date_time.strftime("%d-%m-%y") # ( OPTIONAL : for parking like airports and hotels )
                                        exit_time = date_time.strftime('%H:%M:%S')

                                        sheet2[f'B{entered_slot_number2+1}']=True

                                        for row in range(1,sheet.max_row + 1):
                                                
                                                if sheet.cell(row=row, column=6).value == entered_otp_number2 and sheet.cell(row=row, column=3).value == entered_slot_number2:
                                                    row_number=row
                                                    
                                                    if sheet.cell(row=row, column=7).value==None:
                                                        sheet[f'G{row_number}']=exit_time


                                        exitend_title=tk.Label(window, 
                                                    bg=CYAN, 
                                                    fg=BLACK, 
                                                    text="COMPLETED ✔️",
                                                    font=('georgia',90, "bold")
                                                    )
                                        exitend_title.place(x=280,y=45)

                                        exitend_msg=tk.Label(window,
                                                            background=BGCOLOR,
                                                            foreground=WHITE,
                                                            text='''
Your exit process is completed successfully. Please press the below
confirm button after completing all the process to close this tab.

Thanks for comming and visit again 🙏 !'''     ,
                                                            font=('georgia',33)
                                                            )
                                        exitend_msg.place(x=110,y=250)

                                        def confirm():
                                            confirm_button.destroy()
                                            exitend_title.destroy()
                                            exitend_msg.destroy()
                                            
                                            main_page()

                                        confirm_button=tk.Button(window, 
                                                    background=CYAN,
                                                    foreground=BLACK,
                                                    activebackground=LIGHTCYAN,
                                                    activeforeground=BLACK,
                                                    highlightthickness=2,
                                                    highlightbackground=CYAN,
                                                    highlightcolor=WHITE,
                                                    width=9,
                                                    height=1,
                                                    border=0,
                                                    cursor='hand1',
                                                    text="CONFIRM",
                                                    font=('arial',40,'bold'),
                                                    command=confirm
                                                    )
                                        confirm_button.place(x=620,y=650)

                                        wb.save(user_file)
                                        wb2.save(parking_file)

                            

                            elif  entered_otp_number==entered_slot_number or sheet2[f'B{entered_slot_number2+1}']!=False or sheet2[f'D{entered_slot_number2+1}']!=entered_otp_number2:

                                        wrong_exit_input_title=tk.Label(window, 
                                                    bg=CYAN, 
                                                    fg=BLACK, 
                                                    text="WRONG ❌",
                                                    font=('georgia',90, "bold")
                                                    )
                                        wrong_exit_input_title.place(x=400,y=45)

                                        wrong_exit_input_msg=tk.Label(window,
                                                            background=BGCOLOR,
                                                            foreground=WHITE,
                                                            text='''
The OTP or the slot number that your entered is incorrect
Please enter correct info to complete the entry procedure.

Press the retry button below to return to the previous page!!!

TRY AGAIN !!!'''     ,
                                                            font=('georgia',33)
                                                            )
                                        wrong_exit_input_msg.place(x=175,y=230)

                                        def retry():
                                            retry_button.destroy()
                                            wrong_exit_input_title.destroy()
                                            wrong_exit_input_msg.destroy()
                                            exit_page()

                                        retry_button=tk.Button(window, 
                                                    background=CYAN,
                                                    foreground=BLACK,
                                                    activebackground=LIGHTCYAN,
                                                    activeforeground=BLACK,
                                                    highlightthickness=2,
                                                    highlightbackground=CYAN,
                                                    highlightcolor=WHITE,
                                                    width=9,
                                                    height=1,
                                                    border=0,
                                                    cursor='hand1',
                                                    text="Retry",
                                                    font=('arial',40,'bold'),
                                                    command=retry
                                                    )
                                        retry_button.place(x=620,y=650)

                    
                    except ValueError:
                    # Handle the case where entered_otp_number is not a valid integer
                     # You can display an error message to the user or take any other appropriate action
                            wrong_exit_input_title=tk.Label(window, 
                                            bg=CYAN, 
                                            fg=BLACK, 
                                            text="WRONG ❌",
                                            font=('georgia',90, "bold")
                                            )
                            wrong_exit_input_title.place(x=400,y=45)

                            wrong_exit_input_msg=tk.Label(window,
                                                    background=BGCOLOR,
                                                    foreground=WHITE,
                                                    text='''
The OTP or the slot number that your entered is \"INVALID\"
Please enter correct info to complete the exit procedure.

Press the retry button below to return to the previous page!!!

TRY AGAIN !!!'''     ,
                                                    font=('georgia',33)
                                                    )
                            wrong_exit_input_msg.place(x=175,y=230)

                            def retry():
                                    retry_button.destroy()
                                    wrong_exit_input_title.destroy()
                                    wrong_exit_input_msg.destroy()
                                    exit_page()

                            retry_button=tk.Button(window, 
                                            background=CYAN,
                                            foreground=BLACK,
                                            activebackground=LIGHTCYAN,
                                            activeforeground=BLACK,
                                            highlightthickness=2,
                                            highlightbackground=CYAN,
                                            highlightcolor=WHITE,
                                            width=9,
                                            height=1,
                                            border=0,
                                            cursor='hand1',
                                            text="Retry",
                                            font=('arial',40,'bold'),
                                            command=retry
                                            )
                            retry_button.place(x=620,y=650)
                            
                    
    #--------------- BUTTONS FOR EXIT PAGE STARTS HERE :

            submit_button=tk.Button(window, 
                                background=CYAN,
                                foreground=BLACK,
                                activebackground=LIGHTCYAN,
                                activeforeground=BLACK,
                                highlightthickness=2,
                                highlightbackground=CYAN,
                                highlightcolor=WHITE,
                                width=8,
                                height=1,
                                border=0,
                                cursor='hand1',
                                text="SUBMIT",
                                font=('arial',40,'bold'),
                                command=exit_end,
                                )
            submit_button.place(x=1150,y=350)

            def clear():
                    otp_number.set('')
                    slot_number.set('')

            clear_button=tk.Button(window, 
                                background=CYAN,
                                foreground=BLACK,
                                activebackground=LIGHTCYAN,
                                activeforeground=BLACK,
                                highlightthickness=2,
                                highlightbackground=CYAN,
                                highlightcolor=WHITE,
                                width=8,
                                height=1,
                                border=0,
                                cursor='hand1',
                                text="CLEAR",
                                font=('arial',40,'bold'),
                                command=clear
                                )
            clear_button.place(x=1150,y=500)

            def back():
                    exitpage_title.destroy()
                    exitpage_msg.destroy()
                    submit_button.destroy()
                    clear_button.destroy()
                    back_button.destroy()
                    slotno_title.destroy()
                    otp_title.destroy()
                    otp_number_entry.destroy()
                    slot_number_entry.destroy()
                    main_page()

            back_button=tk.Button(window, 
                                background=CYAN,
                                foreground=BLACK,
                                activebackground=LIGHTCYAN,
                                activeforeground=BLACK,
                                highlightthickness=2,
                                highlightbackground=CYAN,
                                highlightcolor=WHITE,
                                width=8,
                                height=1,
                                border=0,
                                cursor='hand1',
                                text="BACK",
                                font=('arial',40,'bold'),
                                command=back
                                )
            back_button.place(x=1150,y=650)
            
    
        elif total_available_slots==total_slots:
                  
            noslots_filled_title=tk.Label(window, 
                                bg=CYAN, 
                                fg=BLACK, 
                                text="WRONG ❌",
                                font=('georgia',90, "bold")
                                )
            noslots_filled_title.place(x=400,y=45)

            noslots_filled_msg=tk.Label(window,
                                        background=BGCOLOR,
                                        foreground=WHITE,
                                        text='''
Unfortunately, the action you requested cannot be done as there is 
\'NO SLOTS FILLED\'. That is all slots are empty, If your are entering
the parking press \'ENTRY\' button. 

Press the Back button below to return to the previous page!!!

Thank you. Have a nice day. !'''     ,
                                        font=('georgia',27)
                                        )
            noslots_filled_msg.place(x=220,y=250)

            def confirm():
                        noslots_filled_back_button.destroy()
                        noslots_filled_title.destroy()
                        noslots_filled_msg.destroy()
                        main_page()

            noslots_filled_back_button=tk.Button(window, 
                                background=CYAN,
                                foreground=BLACK,
                                activebackground=LIGHTCYAN,
                                activeforeground=BLACK,
                                highlightthickness=2,
                                highlightbackground=CYAN,
                                highlightcolor=WHITE,
                                width=9,
                                height=1,
                                border=0,
                                cursor='hand1',
                                text="BACK",
                                font=('arial',40,'bold'),
                                command=confirm
                                )
            noslots_filled_back_button.place(x=620,y=650)
        

    color_box = tk.Frame(window, 
                        width=1920, 
                        height=220, 
                        background=CYAN,
                        )
    color_box.place(x=0,y=0)

    page_title = tk.Label(window, 
                          bg=CYAN, 
                          fg=BLACK, 
                          text="ADPS PARKING SYSTEM",
                          font=('georgia', 80, "bold")
                          )
    page_title.place(x=60, y=50)

    page_subtitle = tk.Label(window, 
                        background=BGCOLOR,
                        foreground=LIGHTCYAN,
                        text = "Please select the operation : ",
                        font=('georgia',40)
                        )
    page_subtitle.place(x=435,y=250)

    page_msg = tk.Label(window,
                        background=BGCOLOR,
                        foreground=WHITE,
                        text='''
                        ENTRY - For entering the parking.

                        EXIT - For exiting the parking.''',
                        font=('georgia',30)
                        )
    page_msg.place(x=210,y=330)

# BUTTONS FOR MAIN PAGE STARTS HERE :

    entry_button=tk.Button(window, 
                        background=CYAN,
                        foreground=BLACK,
                        activebackground=LIGHTCYAN,
                        activeforeground=BLACK,
                        highlightthickness=2,
                        highlightbackground=CYAN,
                        highlightcolor=WHITE,
                        width=8,
                        height=1,
                        border=0,
                        cursor='hand1',
                        text="ENTRY",
                        font=('arial',50,'bold'),
                        command=entry_page,
                        )
    entry_button.place(x=235,y=600)

    exit_button=tk.Button(window, 
                        background=CYAN,
                        foreground=BLACK,
                        activebackground=LIGHTCYAN,
                        activeforeground=BLACK,
                        highlightthickness=2,
                        highlightbackground=CYAN,
                        highlightcolor=WHITE,
                        width=8,
                        height=1,
                        border=0,
                        cursor='hand1',
                        text=" EXIT ",
                        font=('arial',50,'bold'),
                        command=exit_page
                        )
    exit_button.place(x=955,y=600)
    slots_availability()
    

wb.save(user_file)
wb2.save(parking_file)


main_page()
window.mainloop()